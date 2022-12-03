from tensorflow import keras #神經網路編輯模組
from tensorflow.keras import layers
import time
import numpy as np
import cv2
from openpyxl import load_workbook, workbook
#本程式的功用是用來驗證卷積網路前的正規化是否正確,目的是把陣列變成圖像化
wb = load_workbook('nor.xlsx')
ws = wb.active
t1=0 #列
x=0#欄
t=['A','B','C','D','E','F','G','H',
   'I','J','K','L','M','N','O','P',
   'Q','R','S','T','U','V','W','X',
   'Y','Z','AA','AB']
#----
np.set_printoptions(threshold=np.inf)
img =  cv2.imread('number//0.jpg')
img2_gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY) #轉成黑白

img3 = 255 - img2_gray #黑白調換

img3 = img3.astype('float32')
img3_min = np.amin(img3)#0.0
img4 = img3 - np.amin(img3) #圖片上像素的01
img5 = 255 * img4 / (np.amax(img4)+1)
kernel = np.ones((5,5),np.uint8)
img6 = cv2.dilate(img5,kernel,iterations = 3)
img7 = cv2.resize(img6,(28,28),1)
img8 = img6.astype('uint8')
x_test_image = np.reshape(img7,(1,28,28))
#cv2.imshow('Carmer1',img2_gray)
x_Test4D = x_test_image.reshape(x_test_image.shape[0],28,28,1).astype('float32') #改變形狀
x_Test4D_normalize = ( x_Test4D / np.amax(x_test_image) ) #除以像素最大值
#print(x_Test4D_normalize[0][0])
for i in range (0,27,1):
    for i1 in range (0,27,1):
        a1=t[i1]
        ws[str(a1)+str(i+1)].value = x_Test4D_normalize[0][i][i1][0]
wb.save('nor.xlsx')
print("以儲存")

